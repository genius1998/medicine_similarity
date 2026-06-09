from __future__ import annotations

import hashlib
import hmac
import json
import re
import secrets
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook

from api.config import get_settings
from api.db import default_sqlite_path, sqlite_connection


RUNTIME_CUSTOM_FUNCTIONAL_TABLE_NAME = "runtime_custom_functional_ingredient_map"


def _utcnow() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat()


def _json(value: Any) -> str:
    return json.dumps(value if value is not None else {}, ensure_ascii=False)


def _safe_json_loads(value: Any, default: Any) -> Any:
    if value in (None, ""):
        return default
    try:
        return json.loads(str(value))
    except Exception:  # noqa: BLE001
        return default


@dataclass
class AuthUser:
    user_id: int
    email: str
    full_name: str
    organization: str
    role: str
    is_active: bool


class OperationsService:
    def __init__(self, sqlite_path: Optional[Path] = None) -> None:
        self.sqlite_path = Path(sqlite_path) if sqlite_path else default_sqlite_path()
        self.ensure_tables()
        self.ensure_default_admin()

    def ensure_tables(self) -> None:
        with sqlite_connection(self.sqlite_path) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS app_user (
                    user_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    email TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    full_name TEXT DEFAULT '',
                    organization TEXT DEFAULT '',
                    role TEXT NOT NULL DEFAULT 'user',
                    is_active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS app_session (
                    session_token TEXT PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    expires_at TEXT NOT NULL,
                    last_seen_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS ocr_review_history (
                    review_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    base_trace_id TEXT,
                    latest_trace_id TEXT,
                    user_id INTEGER,
                    source_type TEXT NOT NULL,
                    product_name_candidate TEXT DEFAULT '',
                    ocr_raw_text TEXT DEFAULT '',
                    parsed_json TEXT DEFAULT '{}',
                    normalized_ingredients_json TEXT DEFAULT '[]',
                    edited_ingredients_json TEXT DEFAULT '[]',
                    recommendations_json TEXT DEFAULT '[]',
                    review_status TEXT DEFAULT 'captured',
                    review_notes TEXT DEFAULT '',
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS ingredient_approval_queue (
                    queue_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    functional_ingredient_name TEXT UNIQUE NOT NULL,
                    category_main TEXT DEFAULT '',
                    category_sub TEXT DEFAULT '',
                    function_text TEXT DEFAULT '',
                    claim_text TEXT DEFAULT '',
                    source TEXT DEFAULT '',
                    confidence REAL DEFAULT 0,
                    notes TEXT DEFAULT '',
                    status TEXT DEFAULT 'pending',
                    reviewed_by INTEGER,
                    reviewed_at TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS app_event_log (
                    log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    event_type TEXT NOT NULL,
                    level TEXT NOT NULL DEFAULT 'info',
                    user_id INTEGER,
                    request_path TEXT DEFAULT '',
                    request_method TEXT DEFAULT '',
                    status_code INTEGER,
                    duration_ms REAL DEFAULT 0,
                    message TEXT DEFAULT '',
                    payload_json TEXT DEFAULT '{}',
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_app_user_email ON app_user(email)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_app_session_user_id ON app_session(user_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_ocr_review_history_trace ON ocr_review_history(base_trace_id, latest_trace_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_ingredient_approval_queue_status ON ingredient_approval_queue(status)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_app_event_log_created_at ON app_event_log(created_at)")
            for sql in [
                "ALTER TABLE functional_category_map ADD COLUMN function_text TEXT DEFAULT ''",
                f"ALTER TABLE {RUNTIME_CUSTOM_FUNCTIONAL_TABLE_NAME} ADD COLUMN function_text TEXT DEFAULT ''",
                "ALTER TABLE ingredient_approval_queue ADD COLUMN function_text TEXT DEFAULT ''",
            ]:
                try:
                    conn.execute(sql)
                except sqlite3.OperationalError:
                    pass
            conn.commit()

    def ensure_default_admin(self) -> None:
        settings = get_settings()
        default_email = str(settings.auth_default_admin_email or "").strip().lower()
        default_password = str(settings.auth_default_admin_password or "")
        if not default_email or not default_password:
            return
        with sqlite_connection(self.sqlite_path) as conn:
            existing = conn.execute("SELECT COUNT(*) FROM app_user").fetchone()[0]
            if existing:
                return
        self.create_user(
            email=default_email,
            password=default_password,
            full_name="System Admin",
            organization="SeongbunKok",
            role="admin",
        )

    def _hash_password(self, password: str, salt: Optional[str] = None) -> str:
        salt_value = salt or secrets.token_hex(16)
        digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt_value), 120_000).hex()
        return f"{salt_value}${digest}"

    def _verify_password(self, password: str, stored_hash: str) -> bool:
        salt, _, _digest = str(stored_hash or "").partition("$")
        if not salt or not _digest:
            return False
        candidate = self._hash_password(password, salt)
        return hmac.compare_digest(candidate, stored_hash)

    def create_user(self, *, email: str, password: str, full_name: str = "", organization: str = "", role: str = "user") -> AuthUser:
        normalized_email = str(email or "").strip().lower()
        if not normalized_email:
            raise ValueError("email is required")
        if len(str(password or "")) < 6:
            raise ValueError("password must be at least 6 characters")
        password_hash = self._hash_password(password)
        with sqlite_connection(self.sqlite_path) as conn:
            conn.execute(
                """
                INSERT INTO app_user (email, password_hash, full_name, organization, role, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (normalized_email, password_hash, str(full_name or ""), str(organization or ""), str(role or "user"), _utcnow()),
            )
            conn.commit()
            row = conn.execute(
                "SELECT user_id, email, full_name, organization, role, is_active FROM app_user WHERE email = ?",
                (normalized_email,),
            ).fetchone()
        if not row:
            raise ValueError("failed to create user")
        return AuthUser(
            user_id=int(row[0]),
            email=str(row[1]),
            full_name=str(row[2] or ""),
            organization=str(row[3] or ""),
            role=str(row[4] or "user"),
            is_active=bool(row[5]),
        )

    def authenticate_user(self, email: str, password: str) -> Optional[AuthUser]:
        normalized_email = str(email or "").strip().lower()
        with sqlite_connection(self.sqlite_path) as conn:
            row = conn.execute(
                "SELECT user_id, email, password_hash, full_name, organization, role, is_active FROM app_user WHERE email = ?",
                (normalized_email,),
            ).fetchone()
        if not row or not bool(row[6]) or not self._verify_password(password, str(row[2] or "")):
            return None
        return AuthUser(
            user_id=int(row[0]),
            email=str(row[1]),
            full_name=str(row[3] or ""),
            organization=str(row[4] or ""),
            role=str(row[5] or "user"),
            is_active=bool(row[6]),
        )

    def create_session(self, user_id: int, *, hours: int = 12) -> str:
        token = secrets.token_urlsafe(32)
        expires_at = (datetime.utcnow() + timedelta(hours=hours)).replace(microsecond=0).isoformat()
        with sqlite_connection(self.sqlite_path) as conn:
            conn.execute(
                """
                INSERT INTO app_session (session_token, user_id, expires_at, last_seen_at)
                VALUES (?, ?, ?, ?)
                """,
                (token, int(user_id), expires_at, _utcnow()),
            )
            conn.commit()
        return token

    def delete_session(self, session_token: str) -> None:
        with sqlite_connection(self.sqlite_path) as conn:
            conn.execute("DELETE FROM app_session WHERE session_token = ?", (str(session_token or ""),))
            conn.commit()

    def get_user_by_session(self, session_token: str) -> Optional[AuthUser]:
        cleaned = str(session_token or "").strip()
        if not cleaned:
            return None
        with sqlite_connection(self.sqlite_path) as conn:
            row = conn.execute(
                """
                SELECT u.user_id, u.email, u.full_name, u.organization, u.role, u.is_active, s.expires_at
                FROM app_session s
                JOIN app_user u ON u.user_id = s.user_id
                WHERE s.session_token = ?
                """,
                (cleaned,),
            ).fetchone()
            if not row:
                return None
            expires_at = str(row[6] or "")
            if expires_at and expires_at < _utcnow():
                conn.execute("DELETE FROM app_session WHERE session_token = ?", (cleaned,))
                conn.commit()
                return None
            conn.execute("UPDATE app_session SET last_seen_at = ? WHERE session_token = ?", (_utcnow(), cleaned))
            conn.commit()
        return AuthUser(
            user_id=int(row[0]),
            email=str(row[1]),
            full_name=str(row[2] or ""),
            organization=str(row[3] or ""),
            role=str(row[4] or "user"),
            is_active=bool(row[5]),
        )

    def log_event(
        self,
        *,
        event_type: str,
        level: str = "info",
        user_id: Optional[int] = None,
        request_path: str = "",
        request_method: str = "",
        status_code: Optional[int] = None,
        duration_ms: float = 0.0,
        message: str = "",
        payload: Optional[dict] = None,
    ) -> None:
        with sqlite_connection(self.sqlite_path) as conn:
            conn.execute(
                """
                INSERT INTO app_event_log (
                    event_type, level, user_id, request_path, request_method, status_code,
                    duration_ms, message, payload_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(event_type or ""),
                    str(level or "info"),
                    int(user_id) if user_id is not None else None,
                    str(request_path or ""),
                    str(request_method or ""),
                    int(status_code) if status_code is not None else None,
                    float(duration_ms or 0.0),
                    str(message or ""),
                    _json(payload or {}),
                ),
            )
            conn.commit()

    def save_ocr_review(
        self,
        *,
        user_id: Optional[int],
        source_type: str,
        response_payload: dict,
        edited_ingredients: Optional[List[str]] = None,
        base_trace_id: str = "",
        review_notes: str = "",
    ) -> None:
        parsed = dict(response_payload.get("parsed", {}) or {})
        recommendations = list(response_payload.get("recommendations", []) or [])
        ocr_payload = dict(response_payload.get("ocr", {}) or {})
        normalized_ingredients = list(parsed.get("normalized_ingredients", []) or [])
        trace_id = str(response_payload.get("trace_id", "") or "")
        product_name_candidate = str(parsed.get("product_name_candidate", "") or "")
        base_key = str(base_trace_id or trace_id or "")
        with sqlite_connection(self.sqlite_path) as conn:
            existing = None
            if base_key:
                existing = conn.execute(
                    "SELECT review_id FROM ocr_review_history WHERE base_trace_id = ? OR latest_trace_id = ? ORDER BY review_id DESC LIMIT 1",
                    (base_key, base_key),
                ).fetchone()
            if existing:
                conn.execute(
                    """
                    UPDATE ocr_review_history
                    SET latest_trace_id = ?, user_id = COALESCE(?, user_id), product_name_candidate = ?,
                        ocr_raw_text = ?, parsed_json = ?, normalized_ingredients_json = ?,
                        edited_ingredients_json = ?, recommendations_json = ?, review_status = ?,
                        review_notes = ?, updated_at = ?
                    WHERE review_id = ?
                    """,
                    (
                        trace_id,
                        int(user_id) if user_id is not None else None,
                        product_name_candidate,
                        str(ocr_payload.get("raw_text", "") or ""),
                        _json(parsed),
                        _json(normalized_ingredients),
                        _json(edited_ingredients or []),
                        _json(recommendations[:10]),
                        "edited_recommended" if edited_ingredients else "captured",
                        str(review_notes or ""),
                        _utcnow(),
                        int(existing[0]),
                    ),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO ocr_review_history (
                        base_trace_id, latest_trace_id, user_id, source_type, product_name_candidate,
                        ocr_raw_text, parsed_json, normalized_ingredients_json, edited_ingredients_json,
                        recommendations_json, review_status, review_notes, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        base_key or trace_id,
                        trace_id,
                        int(user_id) if user_id is not None else None,
                        str(source_type or ""),
                        product_name_candidate,
                        str(ocr_payload.get("raw_text", "") or ""),
                        _json(parsed),
                        _json(normalized_ingredients),
                        _json(edited_ingredients or []),
                        _json(recommendations[:10]),
                        "edited_recommended" if edited_ingredients else "captured",
                        str(review_notes or ""),
                        _utcnow(),
                    ),
                )
            conn.commit()

    def list_ocr_reviews(self, *, limit: int = 100) -> List[dict]:
        with sqlite_connection(self.sqlite_path) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                """
                SELECT h.review_id, h.base_trace_id, h.latest_trace_id, h.source_type, h.product_name_candidate,
                       h.review_status, h.review_notes, h.created_at, h.updated_at,
                       h.normalized_ingredients_json, h.edited_ingredients_json, h.recommendations_json,
                       u.email AS user_email
                FROM ocr_review_history h
                LEFT JOIN app_user u ON u.user_id = h.user_id
                ORDER BY h.review_id DESC
                LIMIT ?
                """,
                (int(limit),),
            ).fetchall()
        items = []
        for row in rows:
            items.append(
                {
                    "review_id": int(row["review_id"]),
                    "base_trace_id": str(row["base_trace_id"] or ""),
                    "latest_trace_id": str(row["latest_trace_id"] or ""),
                    "source_type": str(row["source_type"] or ""),
                    "product_name_candidate": str(row["product_name_candidate"] or ""),
                    "review_status": str(row["review_status"] or ""),
                    "review_notes": str(row["review_notes"] or ""),
                    "created_at": str(row["created_at"] or ""),
                    "updated_at": str(row["updated_at"] or ""),
                    "user_email": str(row["user_email"] or ""),
                    "normalized_ingredients": _safe_json_loads(row["normalized_ingredients_json"], []),
                    "edited_ingredients": _safe_json_loads(row["edited_ingredients_json"], []),
                    "recommendations": _safe_json_loads(row["recommendations_json"], []),
                }
            )
        return items

    def sync_pending_ingredients_from_runtime(self) -> None:
        with sqlite_connection(self.sqlite_path) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                f"""
                SELECT functional_ingredient_name, category_main, category_sub, function_text, claim_text, source, confidence, notes
                FROM {RUNTIME_CUSTOM_FUNCTIONAL_TABLE_NAME}
                WHERE source = 'runtime_auto_provisional'
                """
            ).fetchall()
            for row in rows:
                conn.execute(
                    """
                    INSERT INTO ingredient_approval_queue (
                        functional_ingredient_name, category_main, category_sub, function_text, claim_text,
                        source, confidence, notes, status, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)
                    ON CONFLICT(functional_ingredient_name) DO UPDATE SET
                        category_main = excluded.category_main,
                        category_sub = excluded.category_sub,
                        function_text = excluded.function_text,
                        claim_text = excluded.claim_text,
                        source = excluded.source,
                        confidence = excluded.confidence,
                        notes = excluded.notes,
                        updated_at = excluded.updated_at
                    """,
                    (
                        str(row["functional_ingredient_name"] or ""),
                        str(row["category_main"] or ""),
                        str(row["category_sub"] or ""),
                        str(row["function_text"] or ""),
                        str(row["claim_text"] or ""),
                        str(row["source"] or ""),
                        float(row["confidence"] or 0.0),
                        str(row["notes"] or ""),
                        _utcnow(),
                    ),
                )
            conn.commit()

    def list_pending_ingredient_approvals(self, *, limit: int = 200, status: str = "pending") -> List[dict]:
        self.sync_pending_ingredients_from_runtime()
        with sqlite_connection(self.sqlite_path) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                """
                SELECT q.queue_id, q.functional_ingredient_name, q.category_main, q.category_sub,
                       q.function_text, q.claim_text, q.source, q.confidence, q.notes, q.status,
                       q.created_at, q.updated_at, q.reviewed_at, u.email AS reviewed_by_email
                FROM ingredient_approval_queue q
                LEFT JOIN app_user u ON u.user_id = q.reviewed_by
                WHERE q.status = ?
                ORDER BY q.queue_id DESC
                LIMIT ?
                """,
                (str(status or "pending"), int(limit)),
            ).fetchall()
        return [dict(row) for row in rows]

    def review_pending_ingredient(
        self,
        *,
        functional_ingredient_name: str,
        decision: str,
        reviewer_user_id: int,
        category_main: str = "",
        category_sub: str = "",
        function_text: str = "",
        claim_text: str = "",
        notes: str = "",
    ) -> None:
        cleaned_name = str(functional_ingredient_name or "").strip()
        reviewed_at = _utcnow()
        with sqlite_connection(self.sqlite_path) as conn:
            if decision == "approve":
                conn.execute(
                    f"""
                    UPDATE {RUNTIME_CUSTOM_FUNCTIONAL_TABLE_NAME}
                    SET category_main = ?, category_sub = ?, function_text = ?, claim_text = ?, notes = ?, source = 'admin_approved', updated_at = ?
                    WHERE functional_ingredient_name = ?
                    """,
                    (
                        str(category_main or ""),
                        str(category_sub or ""),
                        str(function_text or ""),
                        str(claim_text or ""),
                        str(notes or ""),
                        reviewed_at,
                        cleaned_name,
                    ),
                )
            elif decision == "reject":
                conn.execute(
                    f"DELETE FROM {RUNTIME_CUSTOM_FUNCTIONAL_TABLE_NAME} WHERE functional_ingredient_name = ?",
                    (cleaned_name,),
                )
                conn.execute(
                    """
                    DELETE FROM ingredient_match_cache
                    WHERE matched_standard_name = ?
                      AND match_method IN ('runtime_rag_llm_new_standard', 'runtime_rag_llm_new_standard_self')
                    """,
                    (cleaned_name,),
                )
            else:
                raise ValueError("decision must be approve or reject")
            conn.execute(
                """
                UPDATE ingredient_approval_queue
                SET category_main = ?, category_sub = ?, function_text = ?, claim_text = ?, notes = ?, status = ?, reviewed_by = ?, reviewed_at = ?, updated_at = ?
                WHERE functional_ingredient_name = ?
                """,
                (
                    str(category_main or ""),
                    str(category_sub or ""),
                    str(function_text or ""),
                    str(claim_text or ""),
                    str(notes or ""),
                    "approved" if decision == "approve" else "rejected",
                    int(reviewer_user_id),
                    reviewed_at,
                    reviewed_at,
                    cleaned_name,
                ),
            )
            conn.commit()

    def get_monitoring_snapshot(self) -> dict:
        self.sync_pending_ingredients_from_runtime()
        with sqlite_connection(self.sqlite_path) as conn:
            total_users = conn.execute("SELECT COUNT(*) FROM app_user").fetchone()[0]
            total_reviews = conn.execute("SELECT COUNT(*) FROM ocr_review_history").fetchone()[0]
            pending_approvals = conn.execute("SELECT COUNT(*) FROM ingredient_approval_queue WHERE status = 'pending'").fetchone()[0]
            total_logs = conn.execute("SELECT COUNT(*) FROM app_event_log").fetchone()[0]
            request_rows = conn.execute(
                """
                SELECT COUNT(*) AS cnt, AVG(duration_ms) AS avg_ms
                FROM app_event_log
                WHERE event_type = 'http_request' AND created_at >= datetime('now', '-1 day')
                """
            ).fetchone()
            error_rows = conn.execute(
                """
                SELECT COUNT(*) FROM app_event_log
                WHERE level IN ('error', 'critical') AND created_at >= datetime('now', '-1 day')
                """
            ).fetchone()
        return {
            "total_users": int(total_users or 0),
            "total_reviews": int(total_reviews or 0),
            "pending_approvals": int(pending_approvals or 0),
            "total_logs": int(total_logs or 0),
            "requests_last_day": int((request_rows[0] if request_rows else 0) or 0),
            "avg_request_ms_last_day": round(float((request_rows[1] if request_rows else 0.0) or 0.0), 2),
            "errors_last_day": int((error_rows[0] if error_rows else 0) or 0),
        }

    def count_app_logs(self) -> int:
        with sqlite_connection(self.sqlite_path) as conn:
            row = conn.execute("SELECT COUNT(*) FROM app_event_log").fetchone()
        return int((row[0] if row else 0) or 0)

    def list_app_logs(self, *, limit: int = 200, offset: int = 0) -> List[dict]:
        with sqlite_connection(self.sqlite_path) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                """
                SELECT l.log_id, l.event_type, l.level, l.request_path, l.request_method, l.status_code,
                       l.duration_ms, l.message, l.payload_json, l.created_at, u.email AS user_email
                FROM app_event_log l
                LEFT JOIN app_user u ON u.user_id = l.user_id
                ORDER BY l.log_id DESC
                LIMIT ? OFFSET ?
                """,
                (int(limit), max(0, int(offset))),
            ).fetchall()
        items = []
        for row in rows:
            item = dict(row)
            item["payload"] = _safe_json_loads(item.pop("payload_json", "{}"), {})
            items.append(item)
        return items

    def export_ocr_reviews_xlsx(self) -> bytes:
        rows = self.list_ocr_reviews(limit=1000)
        wb = Workbook()
        ws = wb.active
        ws.title = "ocr_reviews"
        ws.append(["review_id", "user_email", "source_type", "product_name_candidate", "review_status", "created_at", "updated_at", "normalized_ingredients", "edited_ingredients"])
        for row in rows:
            ws.append([
                row["review_id"],
                row["user_email"],
                row["source_type"],
                row["product_name_candidate"],
                row["review_status"],
                row["created_at"],
                row["updated_at"],
                ", ".join(row["normalized_ingredients"]),
                ", ".join(row["edited_ingredients"]),
            ])
        details = wb.create_sheet("recommendations")
        details.append(["review_id", "rank", "product_name", "report_no", "similarity_score", "reason"])
        for row in rows:
            for idx, item in enumerate(row["recommendations"], start=1):
                details.append([
                    row["review_id"],
                    idx,
                    str(item.get("target_product_name") or item.get("product_name") or ""),
                    str(item.get("target_report_no") or item.get("report_no") or ""),
                    float(item.get("similarity_score", 0.0) or 0.0),
                    str(item.get("reason") or ""),
                ])
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def export_pending_ingredients_xlsx(self) -> bytes:
        rows = self.list_pending_ingredient_approvals(limit=2000, status="pending")
        wb = Workbook()
        ws = wb.active
        ws.title = "pending_ingredients"
        ws.append(["functional_ingredient_name", "category_main", "category_sub", "function_text", "claim_text", "confidence", "source", "notes", "created_at"])
        for row in rows:
            ws.append([
                str(row.get("functional_ingredient_name", "")),
                str(row.get("category_main", "")),
                str(row.get("category_sub", "")),
                str(row.get("function_text", "")),
                str(row.get("claim_text", "")),
                float(row.get("confidence", 0.0) or 0.0),
                str(row.get("source", "")),
                str(row.get("notes", "")),
                str(row.get("created_at", "")),
            ])
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def export_app_logs_xlsx(self) -> bytes:
        rows = self.list_app_logs(limit=5000)
        wb = Workbook()
        ws = wb.active
        ws.title = "app_logs"
        ws.append(["log_id", "created_at", "level", "event_type", "user_email", "request_method", "request_path", "status_code", "duration_ms", "message"])
        for row in rows:
            ws.append([
                int(row.get("log_id", 0) or 0),
                str(row.get("created_at", "")),
                str(row.get("level", "")),
                str(row.get("event_type", "")),
                str(row.get("user_email", "")),
                str(row.get("request_method", "")),
                str(row.get("request_path", "")),
                int(row.get("status_code", 0) or 0) if row.get("status_code") is not None else "",
                float(row.get("duration_ms", 0.0) or 0.0),
                str(row.get("message", "")),
            ])
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def list_user_accounts(self, *, limit: int = 500) -> List[dict]:
        with sqlite_connection(self.sqlite_path) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                """
                SELECT
                    u.user_id,
                    u.email,
                    u.full_name,
                    u.organization,
                    u.role,
                    u.is_active,
                    u.created_at,
                    u.updated_at,
                    (
                        SELECT MAX(created_at)
                        FROM app_event_log l
                        WHERE l.user_id = u.user_id
                    ) AS last_activity_at,
                    (
                        SELECT COUNT(*)
                        FROM ocr_review_history h
                        WHERE h.user_id = u.user_id
                    ) AS ocr_review_count,
                    (
                        SELECT COUNT(*)
                        FROM app_event_log l
                        WHERE l.user_id = u.user_id
                          AND (
                              l.event_type = 'product_similarity_lookup'
                              OR (l.event_type = 'http_request' AND l.request_path LIKE '/api/products/%/similar')
                          )
                    ) AS product_analysis_count
                FROM app_user u
                ORDER BY datetime(u.created_at) DESC, u.user_id DESC
                LIMIT ?
                """,
                (int(limit),),
            ).fetchall()
        return [dict(row) for row in rows]

    def list_recent_analysis_activity(self, *, limit: int = 200) -> List[dict]:
        with sqlite_connection(self.sqlite_path) as conn:
            conn.row_factory = sqlite3.Row
            ocr_rows = conn.execute(
                """
                SELECT
                    h.review_id AS activity_id,
                    h.created_at,
                    h.updated_at,
                    h.user_id,
                    u.email AS user_email,
                    h.source_type,
                    h.product_name_candidate,
                    h.normalized_ingredients_json,
                    h.edited_ingredients_json,
                    h.recommendations_json,
                    h.latest_trace_id
                FROM ocr_review_history h
                LEFT JOIN app_user u ON u.user_id = h.user_id
                ORDER BY h.review_id DESC
                LIMIT ?
                """,
                (int(limit),),
            ).fetchall()
            event_rows = conn.execute(
                """
                SELECT
                    l.log_id AS activity_id,
                    l.created_at,
                    l.user_id,
                    u.email AS user_email,
                    l.event_type,
                    l.request_path,
                    l.message,
                    l.payload_json
                FROM app_event_log l
                LEFT JOIN app_user u ON u.user_id = l.user_id
                WHERE l.event_type = 'product_similarity_lookup'
                   OR (l.event_type = 'http_request' AND l.request_path LIKE '/api/products/%/similar')
                ORDER BY l.log_id DESC
                LIMIT ?
                """,
                (int(limit),),
            ).fetchall()

        items: List[dict] = []
        for row in ocr_rows:
            recommendations = _safe_json_loads(row["recommendations_json"], [])
            items.append(
                {
                    "activity_type": f"ocr_{str(row['source_type'] or '')}",
                    "activity_id": int(row["activity_id"]),
                    "created_at": str(row["created_at"] or ""),
                    "updated_at": str(row["updated_at"] or ""),
                    "user_id": int(row["user_id"]) if row["user_id"] is not None else None,
                    "user_email": str(row["user_email"] or ""),
                    "source_label": f"OCR/{str(row['source_type'] or '').upper()}",
                    "analyzed_target": str(row["product_name_candidate"] or ""),
                    "analyzed_report_no": "",
                    "trace_id": str(row["latest_trace_id"] or ""),
                    "normalized_ingredients": _safe_json_loads(row["normalized_ingredients_json"], []),
                    "edited_ingredients": _safe_json_loads(row["edited_ingredients_json"], []),
                    "top_results": [
                        {
                            "product_name": str(item.get("target_product_name") or item.get("product_name") or ""),
                            "report_no": str(item.get("target_report_no") or item.get("report_no") or ""),
                            "similarity_score": float(item.get("similarity_score", 0.0) or 0.0),
                        }
                        for item in recommendations[:3]
                    ],
                    "result_recorded": True,
                }
            )

        report_no_pattern = re.compile(r"/api/products/([^/]+)/similar")
        db_event_items: List[dict] = []
        for row in event_rows:
            payload = _safe_json_loads(row["payload_json"], {})
            request_path = str(row["request_path"] or "")
            matched = report_no_pattern.search(request_path)
            report_no = str(payload.get("report_no") or (matched.group(1) if matched else "") or "")
            top_results = []
            for item in list(payload.get("top_results", []) or [])[:3]:
                top_results.append(
                    {
                        "product_name": str(item.get("product_name") or ""),
                        "report_no": str(item.get("report_no") or ""),
                        "similarity_score": float(item.get("similarity_score", 0.0) or 0.0),
                    }
                )
            db_event_items.append(
                {
                    "activity_type": "db_product",
                    "activity_id": int(row["activity_id"]),
                    "created_at": str(row["created_at"] or ""),
                    "updated_at": str(row["created_at"] or ""),
                    "user_id": int(row["user_id"]) if row["user_id"] is not None else None,
                    "user_email": str(row["user_email"] or ""),
                    "source_label": "기존DB 제품",
                    "analyzed_target": str(payload.get("base_product_name") or ""),
                    "analyzed_report_no": report_no,
                    "trace_id": "",
                    "normalized_ingredients": [],
                    "edited_ingredients": [],
                    "top_results": top_results,
                    "llm_rerank_applied": bool(payload.get("llm_rerank_applied", False)),
                    "result_recorded": bool(top_results),
                    "message": str(row["message"] or ""),
                    "request_path": request_path,
                    "event_type": str(row["event_type"] or ""),
                }
            )

        deduped_db_events: Dict[tuple, dict] = {}
        for item in db_event_items:
            dedupe_key = (
                item.get("user_id"),
                item.get("analyzed_report_no"),
                item.get("created_at"),
            )
            existing = deduped_db_events.get(dedupe_key)
            if existing is None:
                deduped_db_events[dedupe_key] = item
                continue
            existing_score = (
                1 if existing.get("event_type") == "product_similarity_lookup" else 0,
                1 if existing.get("result_recorded") else 0,
                int(existing.get("activity_id", 0) or 0),
            )
            candidate_score = (
                1 if item.get("event_type") == "product_similarity_lookup" else 0,
                1 if item.get("result_recorded") else 0,
                int(item.get("activity_id", 0) or 0),
            )
            if candidate_score > existing_score:
                deduped_db_events[dedupe_key] = item

        for item in deduped_db_events.values():
            item.pop("event_type", None)
            items.append(item)

        items.sort(key=lambda item: (str(item.get("created_at", "")), int(item.get("activity_id", 0))), reverse=True)
        return items[: max(1, int(limit))]

    def list_ingredient_catalog(self, *, q: str = "", origin: str = "all", limit: int = 300) -> List[dict]:
        return self.list_ingredient_catalog_page(q=q, origin=origin, page=1, page_size=limit)["results"]

    def list_ingredient_catalog_page(
        self,
        *,
        q: str = "",
        origin: str = "all",
        page: int = 1,
        page_size: int = 50,
        sync_pending: bool = True,
    ) -> dict:
        if sync_pending:
            self.sync_pending_ingredients_from_runtime()
        query = str(q or "").strip()
        origin_filter = str(origin or "all").strip().lower()
        if origin_filter not in {"all", "existing", "new"}:
            origin_filter = "all"
        safe_page = max(1, int(page or 1))
        safe_page_size = max(1, min(1000, int(page_size or 50)))

        combined_sql = f"""
            SELECT f.functional_ingredient_name, f.category_main, f.category_sub, f.function_text, f.claim_text,
                   f.source, f.confidence, f.notes, '' AS created_at, '' AS updated_at,
                   'existing' AS origin_type, '기존 원료' AS origin_label, 'functional_category_map' AS source_table,
                   1 AS origin_rank
            FROM functional_category_map f
            WHERE NOT EXISTS (
                SELECT 1
                FROM {RUNTIME_CUSTOM_FUNCTIONAL_TABLE_NAME} c
                WHERE c.functional_ingredient_name = f.functional_ingredient_name
            )
            UNION ALL
            SELECT c.functional_ingredient_name, c.category_main, c.category_sub, c.function_text, c.claim_text,
                   c.source, c.confidence, c.notes, c.created_at, c.updated_at,
                   'new' AS origin_type, '신규 추가' AS origin_label, '{RUNTIME_CUSTOM_FUNCTIONAL_TABLE_NAME}' AS source_table,
                   0 AS origin_rank
            FROM {RUNTIME_CUSTOM_FUNCTIONAL_TABLE_NAME} c
        """
        where_parts: List[str] = []
        params: List[Any] = []
        if origin_filter in {"existing", "new"}:
            where_parts.append("origin_type = ?")
            params.append(origin_filter)
        if query:
            like = f"%{query}%"
            where_parts.append(
                """
                (
                    functional_ingredient_name LIKE ?
                    OR category_main LIKE ?
                    OR category_sub LIKE ?
                    OR function_text LIKE ?
                    OR claim_text LIKE ?
                    OR notes LIKE ?
                    OR source LIKE ?
                )
                """
            )
            params.extend([like] * 7)
        where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
        with sqlite_connection(self.sqlite_path) as conn:
            conn.row_factory = sqlite3.Row
            total_count = int(
                conn.execute(
                    f"WITH combined AS ({combined_sql}) SELECT COUNT(*) FROM combined {where_sql}",
                    params,
                ).fetchone()[0]
                or 0
            )
            total_pages = max(1, (total_count + safe_page_size - 1) // safe_page_size)
            current_page = min(safe_page, total_pages)
            offset = (current_page - 1) * safe_page_size
            rows = conn.execute(
                f"""
                WITH combined AS ({combined_sql})
                SELECT functional_ingredient_name, category_main, category_sub, function_text, claim_text,
                       source, confidence, notes, created_at, updated_at, origin_type, origin_label, source_table
                FROM combined
                {where_sql}
                ORDER BY origin_rank ASC, functional_ingredient_name ASC
                LIMIT ? OFFSET ?
                """,
                [*params, safe_page_size, offset],
            ).fetchall()
        results: List[dict] = []
        for row in rows:
            item = dict(row)
            item["functional_ingredient_name"] = str(item.get("functional_ingredient_name", "") or "")
            item["category_main"] = str(item.get("category_main", "") or "")
            item["category_sub"] = str(item.get("category_sub", "") or "")
            item["function_text"] = str(item.get("function_text", "") or "")
            item["claim_text"] = str(item.get("claim_text", "") or "")
            item["source"] = str(item.get("source", "") or "")
            item["confidence"] = float(item.get("confidence", 0.0) or 0.0)
            item["notes"] = str(item.get("notes", "") or "")
            item["created_at"] = str(item.get("created_at", "") or "")
            item["updated_at"] = str(item.get("updated_at", "") or "")
            results.append(item)
        return {
            "query": query,
            "origin": origin_filter,
            "page": current_page,
            "page_size": safe_page_size,
            "total_count": total_count,
            "total_pages": total_pages,
            "has_prev": current_page > 1,
            "has_next": current_page < total_pages,
            "prev_page": max(1, current_page - 1),
            "next_page": min(total_pages, current_page + 1),
            "results": results,
        }

    def update_ingredient_catalog_item(
        self,
        *,
        functional_ingredient_name: str,
        source_table: str,
        category_main: str = "",
        category_sub: str = "",
        function_text: str = "",
        claim_text: str = "",
        notes: str = "",
    ) -> None:
        cleaned_name = str(functional_ingredient_name or "").strip()
        cleaned_table = str(source_table or "").strip()
        if cleaned_table not in {"functional_category_map", RUNTIME_CUSTOM_FUNCTIONAL_TABLE_NAME}:
            raise ValueError("invalid source_table")
        with sqlite_connection(self.sqlite_path) as conn:
            if cleaned_table == "functional_category_map":
                conn.execute(
                    """
                    UPDATE functional_category_map
                    SET category_main = ?, category_sub = ?, function_text = ?, claim_text = ?, notes = ?
                    WHERE functional_ingredient_name = ?
                    """,
                    (
                        str(category_main or ""),
                        str(category_sub or ""),
                        str(function_text or ""),
                        str(claim_text or ""),
                        str(notes or ""),
                        cleaned_name,
                    ),
                )
            else:
                conn.execute(
                    f"""
                    UPDATE {RUNTIME_CUSTOM_FUNCTIONAL_TABLE_NAME}
                    SET category_main = ?, category_sub = ?, function_text = ?, claim_text = ?, notes = ?, updated_at = ?
                    WHERE functional_ingredient_name = ?
                    """,
                    (
                        str(category_main or ""),
                        str(category_sub or ""),
                        str(function_text or ""),
                        str(claim_text or ""),
                        str(notes or ""),
                        _utcnow(),
                        cleaned_name,
                    ),
                )
                conn.execute(
                    """
                    UPDATE ingredient_approval_queue
                    SET category_main = ?, category_sub = ?, function_text = ?, claim_text = ?, notes = ?, updated_at = ?
                    WHERE functional_ingredient_name = ?
                    """,
                    (
                        str(category_main or ""),
                        str(category_sub or ""),
                        str(function_text or ""),
                        str(claim_text or ""),
                        str(notes or ""),
                        _utcnow(),
                        cleaned_name,
                    ),
                )
            conn.commit()


_ops_service_singleton: Optional[OperationsService] = None


def get_ops_service() -> OperationsService:
    global _ops_service_singleton
    if _ops_service_singleton is None:
        _ops_service_singleton = OperationsService()
    return _ops_service_singleton
